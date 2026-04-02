import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import (
    carregar_clientes, carregar_etapas, carregar_implantacao,
    inserir, registrar_historico, normalizar_doc, CHECKLIST_SEP
)
from utils import somente_numeros

st.set_page_config(layout="wide")
st.title("📋 Pré-Implantação de Clientes")

# ==============================
# FUNÇÕES AUXILIARES
# ==============================

def limpar_documento(valor):
    if pd.isna(valor):
        return ""
    return somente_numeros(str(valor))

def limpar_checkboxes():
    for k in list(st.session_state.keys()):
        if k.startswith("etapa_"):
            del st.session_state[k]

def gerar_excel_cliente(cpf, cliente, resp_medicit, resp_cliente, etapas_sel, etapas_grouped):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Implantacao"

    headers = [
        "Chave", "cnpj_cpf", "Cliente", "Etapa", "Checklist",
        "Responsavel_Medicit", "Responsavel_Clinica", "Participantes", "Data_Prevista"
    ]
    ws.append(headers)

    header_fill    = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    editable_fill  = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    preench_fill   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for etapa in etapas_sel:
        itens_row = etapas_grouped.loc[etapas_grouped["etapa"] == etapa, "item"]
        itens     = itens_row.values[0] if len(itens_row) > 0 else []
        chave     = f"{cpf}||{etapa}"
        ws.append([
            chave, str(cpf), cliente, etapa,
            " | ".join(itens),
            resp_medicit,
            resp_cliente,   # pré-preenchido do cadastro, destacado em verde
            "",
            ""
        ])

    # Destaca coluna Responsavel_Clinica (col 7) em verde se já preenchida
    for row in range(2, ws.max_row + 1):
        val_clinica = ws.cell(row=row, column=7).value
        for col in [8, 9]:    # Participantes, Data_Prevista — editáveis (amarelo)
            ws.cell(row=row, column=col).fill = editable_fill
        if val_clinica:       # Responsavel_Clinica preenchido — verde claro
            ws.cell(row=row, column=7).fill = preench_fill
            ws.cell(row=row, column=7).font = Font(italic=True)
        else:                 # vazio — amarelo para preencher
            ws.cell(row=row, column=7).fill = editable_fill

    ws.freeze_panes = "A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    return wb

# ==============================
# DADOS
# ==============================

clientes_df    = carregar_clientes()
etapas_df      = carregar_etapas()     # já vem ordenado
implantacao_df = carregar_implantacao()

for df_obj in [clientes_df, etapas_df, implantacao_df]:
    df_obj.columns = df_obj.columns.str.strip().str.lower()

for col in ["cnpj_cpf", "nome", "responsavel_medicit", "responsavel_cliente"]:
    if col not in clientes_df.columns:
        clientes_df[col] = ""

for col in ["cnpj_cpf", "etapa", "chave"]:
    if col not in implantacao_df.columns:
        implantacao_df[col] = ""

clientes_df["cnpj_cpf"]    = clientes_df["cnpj_cpf"].apply(normalizar_doc)
implantacao_df["cnpj_cpf"] = implantacao_df["cnpj_cpf"].apply(normalizar_doc)

clientes_df    = clientes_df[clientes_df["cnpj_cpf"] != ""]
implantacao_df = implantacao_df[implantacao_df["cnpj_cpf"] != ""]

if not etapas_df.empty:
    etapas_df = etapas_df.sort_values(["ordem_etapa", "ordem_item"])

etapas_grouped = (
    etapas_df.groupby(["ordem_etapa", "etapa"], as_index=False)
    .agg({"item": lambda x: [i for i in x if str(i).strip() != ""]})
    .sort_values("ordem_etapa")
)

# ==============================
# SESSION STATE
# ==============================

if "dados_excel"   not in st.session_state:
    st.session_state["dados_excel"] = None
if "cliente_atual" not in st.session_state:
    st.session_state["cliente_atual"] = None

# ==============================
# BUSCA
# ==============================

st.subheader("Buscar Cliente")
busca = st.text_input("Digite CPF/CNPJ ou Nome")

if busca:
    busca_limpa = normalizar_doc(busca)
    filtrados = clientes_df[
        clientes_df["nome"].str.contains(busca, case=False, na=False) |
        clientes_df["cnpj_cpf"].str.contains(busca_limpa, na=False)
    ]
else:
    filtrados = clientes_df

if filtrados.empty:
    st.warning("Nenhum cliente encontrado.")
    st.stop()

cliente_map  = dict(zip(filtrados["nome"], filtrados["cnpj_cpf"]))
cliente_nome = st.selectbox("Selecione o Cliente", list(cliente_map.keys()))

cnpj_cpf = normalizar_doc(cliente_map.get(cliente_nome, ""))
if not cnpj_cpf:
    st.error("Cliente sem CPF/CNPJ válido.")
    st.stop()

# ==============================
# RESPONSÁVEL (vem do cadastro)
# ==============================

dados_cli       = clientes_df[clientes_df["cnpj_cpf"] == cnpj_cpf]
resp_med_padrao = dados_cli.iloc[0].get("responsavel_medicit", "") if not dados_cli.empty else ""
resp_cli_padrao = dados_cli.iloc[0].get("responsavel_cliente", "") if not dados_cli.empty else ""

resp_medicit = st.text_input("Responsável Medicit", value=resp_med_padrao)

if resp_cli_padrao:
    st.info(f"ℹ️ Responsável cliente do cadastro: **{resp_cli_padrao}** — confirme ou altere abaixo.")
resp_cliente = st.text_input(
    "Responsável Cliente (clínica)",
    value=resp_cli_padrao,
    help="Pré-preenchido do cadastro do cliente. Altere se necessário."
)

# Reset ao trocar de cliente
if st.session_state["cliente_atual"] != cnpj_cpf:
    st.session_state["cliente_atual"] = cnpj_cpf
    limpar_checkboxes()

# ==============================
# SELEÇÃO DE ETAPAS
# ==============================

st.subheader("Seleção de Etapas")
etapas_existentes = implantacao_df[implantacao_df["cnpj_cpf"] == cnpj_cpf]["etapa"].tolist()

colunas  = st.columns(3)
selecoes = {}

for i, row in etapas_grouped.iterrows():
    col      = colunas[i % 3]
    key      = f"etapa_{cnpj_cpf}_{row['etapa']}"
    ja_existe = row["etapa"] in etapas_existentes

    if key not in st.session_state:
        st.session_state[key] = False

    with col:
        st.markdown(f"**{'✅ ' if ja_existe else ''}{row['etapa']}**")
        selecoes[row["etapa"]] = st.checkbox(
            "Selecionar" + (" (já cadastrada)" if ja_existe else ""),
            key=key,
            disabled=ja_existe
        )
        for item in row["item"]:
            st.markdown(f"- {item}")

selecionadas = [e for e, v in selecoes.items() if v]

st.subheader("Etapas Selecionadas")
if selecionadas:
    for e in selecionadas:
        st.write(f"• {e}")
else:
    st.write("Nenhuma etapa nova selecionada.")

# ==============================
# SALVAR
# ==============================

with st.form("form_salvar"):
    submit = st.form_submit_button("💾 Salvar Pré-Implantação")

    if submit:
        if not resp_medicit.strip():
            st.error("Informe o Responsável Medicit.")
            st.stop()
        if not selecionadas:
            st.error("Selecione ao menos uma etapa nova.")
            st.stop()

        novas = []
        for etapa in selecionadas:
            itens_row = etapas_grouped.loc[etapas_grouped["etapa"] == etapa, "item"]
            itens     = itens_row.values[0] if len(itens_row) > 0 else []
            chave     = f"{cnpj_cpf}||{etapa}"

            registro = {
                "chave":               chave,
                "cnpj_cpf":            str(cnpj_cpf),
                "cliente":             cliente_nome,
                "etapa":               etapa,
                "status":              "Não iniciado",
                "data_inicio":         datetime.now().strftime("%d/%m/%Y"),
                "checklist":           CHECKLIST_SEP.join(itens),
                "responsavel_medicit": resp_medicit,
                "responsavel_cliente": resp_cliente,
                "participantes":       "",
                "motivo":              "",
                "proxima_acao":        "",
                "ultima_atualizacao":  datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "data_conclusao":      ""
            }
            inserir("implantacao", registro)
            novas.append(etapa)

        if novas:
            registrar_historico(
                cnpj_cpf=cnpj_cpf, cliente=cliente_nome,
                pagina="Pré-Implantação", acao="Inserção de etapas",
                etapa=", ".join(novas)
            )
            st.session_state["dados_excel"] = {
                "cpf":          cnpj_cpf,
                "cliente":      cliente_nome,
                "etapas":       selecionadas,
                "resp_medicit": resp_medicit,
                "resp_cliente": resp_cliente
            }
            st.success(f"✅ {len(novas)} etapa(s) cadastrada(s) com sucesso!")
            limpar_checkboxes()

# ==============================
# DOWNLOAD EXCEL
# ==============================

if st.session_state["dados_excel"]:
    if st.button("📤 Gerar Planilha de Retorno"):
        d  = st.session_state["dados_excel"]
        wb = gerar_excel_cliente(
            d["cpf"], d["cliente"],
            d["resp_medicit"], d["resp_cliente"],
            d["etapas"], etapas_grouped
        )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            label="⬇️ Baixar Excel",
            data=buf,
            file_name=f"pre_implantacao_{d['cpf']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )