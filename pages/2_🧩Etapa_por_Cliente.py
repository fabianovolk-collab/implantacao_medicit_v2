import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# =============================
# Database
# =============================
from database import (
    carregar_clientes,
    carregar_etapas,
    carregar_implantacao,
    salvar_implantacao
)

# =============================
# Funções Auxiliares
# =============================
def limpar_checkboxes():
    keys_para_remover = [k for k in st.session_state.keys() if k.startswith("etapa_")]
    for k in keys_para_remover:
        del st.session_state[k]

def gerar_excel_cliente(cpf, cliente, responsavel_medicit, etapas_sel, etapas_grouped):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Implantacao"

    headers = [
        "CNPJ_CPF","Cliente","Etapa","Checklist",
        "Responsavel_Medicit","Responsavel_Clinica","Participantes","Data_Prevista"
    ]

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    for etapa in etapas_sel:
        itens = etapas_grouped.loc[etapas_grouped["Etapa"] == etapa, "Item"].values[0]
        ws.append([
            cpf,
            cliente,
            etapa,
            " | ".join(itens),
            responsavel_medicit,
            "",
            "",
            ""
        ])

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=6).fill = fill
        ws.cell(row=row, column=7).fill = fill
        ws.cell(row=row, column=8).fill = fill

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    return wb

# =============================
# Carregar dados
# =============================
clientes_df = carregar_clientes()
etapas_df = carregar_etapas()
etapas_df = etapas_df.sort_values(["Ordem_etapa", "Ordem_item"])

# Agrupamento de etapas com lista de itens
etapas_grouped = (
    etapas_df.groupby(["Ordem_etapa", "Etapa"], as_index=False)
    .agg({"Item": lambda x: [i for i in x if str(i).strip() != ""]})
)

implantacao_df = carregar_implantacao()

# =============================
# Session State
# =============================
if "confirmar_substituicao" not in st.session_state:
    st.session_state["confirmar_substituicao"] = False
if "dados_excel" not in st.session_state:
    st.session_state["dados_excel"] = None
if "cliente_anterior" not in st.session_state:
    st.session_state["cliente_anterior"] = ""
if "cliente_existe_implantacao" not in st.session_state:
    st.session_state["cliente_existe_implantacao"] = False

# =============================
# UI
# =============================
st.set_page_config(layout="wide")
st.title("📋 Pré-Implantação de Clientes")

# =============================
# Buscar cliente
# =============================
st.subheader("Buscar Cliente")
busca = st.text_input("Digite CPF/CNPJ ou Nome")

if busca:
    clientes_filtrados = clientes_df[
        clientes_df["Nome"].str.contains(busca, case=False, na=False) |
        clientes_df["CNPJ_CPF"].str.contains(busca, case=False, na=False)
    ]
else:
    clientes_filtrados = clientes_df

cliente_selecionado = st.selectbox(
    "Selecione o Cliente",
    clientes_filtrados["Nome"].tolist() if not clientes_filtrados.empty else ["Nenhum cliente encontrado"]
)

responsavel_medicit = st.text_input("Responsável Medicit")

# Reset ao trocar cliente
if cliente_selecionado != st.session_state["cliente_anterior"]:
    st.session_state["cliente_anterior"] = cliente_selecionado
    limpar_checkboxes()
    st.rerun()

cpf_cnpj_cliente = ""
if cliente_selecionado != "Nenhum cliente encontrado":
    cpf_cnpj_cliente = clientes_filtrados.loc[
        clientes_filtrados["Nome"] == cliente_selecionado, "CNPJ_CPF"
    ].values[0]
else:
    limpar_checkboxes()

# =============================
# Etapas
# =============================
st.subheader("Seleção de Etapas")

colunas = st.columns(3)
selecoes = {}

for i, row in etapas_grouped.iterrows():
    col = colunas[i % 3]
    key = f"etapa_{cpf_cnpj_cliente}_{row['Etapa']}"
    if key not in st.session_state:
        st.session_state[key] = False

    bg = "#b3e6b3" if st.session_state[key] else "#f0f8ff"

    with col:
        st.markdown(
            f"<div style='background:{bg};padding:8px;border-radius:8px;margin-bottom:8px;border:1px solid #ccc'>",
            unsafe_allow_html=True
        )

        c1, c2 = st.columns([0.15, 0.85])
        with c1:
            selecoes[row["Etapa"]] = st.checkbox("", key=key)
        with c2:
            st.markdown(f"### {row['Etapa']}")

        for item in row["Item"]:
            st.markdown(
                f"<div style='font-size:15px;margin:2px 0px;display:flex;align-items:center;'>"
                f"<span style='margin-right:6px;'>✔️</span>"
                f"<span>{item}</span></div>",
                unsafe_allow_html=True
            )
        st.markdown("</div>", unsafe_allow_html=True)

# =============================
# Etapas selecionadas
# =============================
st.subheader("Etapas Selecionadas")
selecionadas = [e for e, v in selecoes.items() if v]
st.write(selecionadas if selecionadas else "Nenhuma etapa selecionada")

# =============================
# Salvar Pré-Implantação
# =============================
if st.button("Salvar Pré-Implantação"):
    if cliente_selecionado == "Nenhum cliente encontrado":
        st.error("Selecione um cliente válido.")
        st.stop()
    if not responsavel_medicit.strip():
        st.error("Informe o Responsável Medicit.")
        st.stop()
    if not selecionadas:
        st.error("Selecione pelo menos uma etapa.")
        st.stop()

    cliente_existe = not implantacao_df[implantacao_df["CNPJ_CPF"] == cpf_cnpj_cliente].empty
    st.session_state["cliente_existe_implantacao"] = cliente_existe
    st.session_state["confirmar_substituicao"] = True

# =============================
# Confirmar substituição ou adição
# =============================
if st.session_state.get("confirmar_substituicao", False):
    cliente_existe = st.session_state.get("cliente_existe_implantacao", False)
    opcao = "Adicionar"

    if cliente_existe:
        st.warning("⚠️ Cliente já possui implantação cadastrada.")
        opcao = st.radio(
            "Escolha como deseja proceder:",
            ["Apenas adicionar novas etapas", "Substituir tudo"],
            index=0
        )
    else:
        st.info("Nova implantação será criada.")

    if st.button("Confirmar e Salvar"):
        data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df_impl = carregar_implantacao()

        if cliente_existe and opcao == "Substituir tudo":
            df_impl = df_impl[df_impl["CNPJ_CPF"] != cpf_cnpj_cliente]
            etapas_existentes = []
        else:
            df_cliente = df_impl[df_impl["CNPJ_CPF"] == cpf_cnpj_cliente]
            etapas_existentes = df_cliente["Etapa"].tolist()

        novas = []
        for etapa in selecionadas:
            if etapa in etapas_existentes:
                continue
            itens = etapas_grouped.loc[etapas_grouped["Etapa"] == etapa, "Item"].values[0]
            novas.append({
                "CNPJ_CPF": cpf_cnpj_cliente,
                "Cliente": cliente_selecionado,
                "Etapa": etapa,
                "Status": "Pré-implantação",
                "Data_Inicio": data,
                "Data_Conclusao": "",
                "Responsavel_Etapa": "",
                "Proxima_Acao": "",
                "Checklist": "; ".join(itens),
                "Ultima_Atualizacao": data,
                "Motivo": "",
                "Responsavel_Medicit": responsavel_medicit,
                "Responsavel_Clinica": "",
                "Observacao": "",
                "Responsavel_Cliente": "",
                "Participantes": ""
            })

        if novas:
            df_impl = pd.concat([df_impl, pd.DataFrame(novas)], ignore_index=True)

        salvar_implantacao(df_impl)

        st.session_state["dados_excel"] = {
            "cpf": cpf_cnpj_cliente,
            "cliente": cliente_selecionado,
            "etapas": selecionadas,
            "responsavel_medicit": responsavel_medicit
        }

        st.success("Pré-implantação salva com sucesso!")
        st.session_state["confirmar_substituicao"] = False
        limpar_checkboxes()
        st.rerun()

# =============================
# Download Excel
# =============================
if st.session_state["dados_excel"]:
    if st.button("📤 Gerar Planilha para Cliente"):
        dados = st.session_state["dados_excel"]
        wb = gerar_excel_cliente(
            dados["cpf"],
            dados["cliente"],
            dados["responsavel_medicit"],
            dados["etapas"],
            etapas_grouped
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        arquivo = f"pre_implantacao_{dados['cpf']}.xlsx"

        st.download_button(
            label="⬇️ Baixar Planilha",
            data=buffer,
            file_name=arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )